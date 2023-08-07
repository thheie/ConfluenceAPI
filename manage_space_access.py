import json
import requests
import openpyxl
from requests.auth import HTTPBasicAuth
from csv import DictReader
from api_decode_proc import find_permission_type
from api_decode_proc import exel_collumn_mapping
from logger import logger


def readFromCsv(filename):
    with open(filename, 'r') as f:
        dict_reader = DictReader(f)
        list_of_dict = list(dict_reader)
        # print(list_of_dict)
    return list_of_dict

class Confluence_api():

    def __init__(self):
        '''
        Initializes the Confluence_API object
        '''
        self.space_dict = {}
        self.space_user_dict = {}
        self.space_permission_list_raw = []
        self.space_permission_list=[]
        self.space_permissision_tab=[]
        setup_params_list = readFromCsv('setup.csv')
        setup_params = setup_params_list[0]
        self.base_url = setup_params['baseurl']
        self.username = setup_params['username']
        self.token = setup_params['token']
        self.auth = HTTPBasicAuth(self.username, self.token)
        logger.info('Script initialiced')

    def set_space_permission(self, space_key, identifier, type, key, target):
        '''
        Sets a specific permission to a confluence space for a user, or group. All permissions must be set separately.
        Valid permissions are:
            'create': 'page', 'blogpost', 'comment', 'attachment'
            'read': 'space'
            'delete': 'page', 'blogpost', 'comment', 'attachment', 'space'
            'export': 'space'
            'administer': 'space'
            'archive': 'page'
            'restrict_content': 'space'
        :param space_key: the key used by Confluence to identefy the space
        :param identifier: name off the group or email of the user that gets this permission
        :param type: user or group
        :param key: What type off persmission that is given Valid values are: Create, read, delete, export, archive, restrict_content
        :param target: What type of object the persmission is for: space, page, blogpost, comment, attachment
        :return: Api responce.status_code
        '''
        #todo endre slik at metoden også hånterer personer grupper og apper
        url = f'https://{self.base_url}/wiki/rest/api/space/{space_key}/permission'
        headers = {
                   "Accept": "application/json",
                   "Content-Type": "application/json"
                    }
        payload = json.dumps({
                              "subject": {
                                            "type":type,
                                            "identifier":identifier
                                            },
                              "operation": {
                                            "key": key,
                                            "target": target
                                            },
                              "_links": {}
                                })
        # Send API-forespørselen
        response = requests.request(
                                 "POST",
                                 url,
                                 data=payload,
                                 auth=self.auth,
                                 headers=headers)
        # Sjekk svaret fra serveren
        if response.status_code == 200:
            logger.info(f'Lesetilgang er gitt til {key}, {identifier}')
            # print(f'Lesetilgang er gitt til {key}, {identifier}')
        #     todo Oppdaterer space_permission_list (pri: lav)
        else:
            logger.info(f'En feil oppsto: Statuskode: {response.status_code}')
            print(f'En feil oppsto: Statuskode: {response.status_code}')
            print(f'Feilmelding: {response.text}')
        return response.status_code
    def delete_space_permission(self, space_key, id):
        '''
        Deleite a spesicic space permission
        :param space_key:  the key used by Confluence to identefy the space
        :param id: the id of the mermission that is to be deleted
        :return: Status code from the API (204 is ok)
        '''
        url = f'https://{self.base_url}/wiki/rest/api/space/{space_key}/permission/{id}'
        response = requests.request(
            "DELETE",
            url,
            auth=self.auth
        )
        if response.status_code == 204:
            logger.debug(f'The permission {id} to {space_key} has been deleted')
        else:
            logger.critical(f'Failed to delete permission {id} in {space_key}')
            logger.critical(f'{response.text}')
        return response.status_code




    def get_space_permissions(self,load_space_list=[]):
        '''
        Loads all permissions for the spaces listed in the file "confluence_space.csv"
        :param load_space_list: a list of spaces that will be process. If empty all spaces are processed.
        :return:
        '''
        space_loaded = 0
        space_list = readFromCsv('confluence_space.csv')
        for space in space_list:
            if space['space_key'] in load_space_list or len(space_list) == 0:
                spacekey = space['space_key']
                url =  f'https://{self.base_url}/wiki/rest/api/space/{spacekey}?expand=permissions'
                headers = {'Accept': 'application/json'}
                response = requests.request(
                    'GET',
                    url,
                    headers=headers,
                    auth=self.auth
                    )

                if response.status_code == 200:
                    logger.info(f'Loaded all permissions for space: {space["space_name"]}')
                    print(f'Loaded all permissions for space: {space["space_name"]}')
                    result = json.loads(response.text)
                    self.space_permission_list_raw.append(result)
                    space_loaded +=1
                else:
                    logtext = f'Error loading space permissions for space {space["space_name"]}. Status code:,{response.status_code}'
                    print(logtext)
                    logger.critical(logtext)
                    logger.critical(f'{response.text}')
                    # load_permission_sucess = False
        logtext = f'Loaded {space_loaded} of {len(space_list)} spaces'
        print (logtext)
        logger.info(logtext)

    def decode_space_persmissions(self, space_list=[]):
        '''
        Decode the contence of space_permission_list_raw and stores it in space_permission_list. The list space_dict
        (a dict containing all spaces that are loaded) and space_user_dict a dict containin infomation about the
        users/grups/apps that have permissions in one or more spaces listed in space_permission_list_raw.
        :param space_list: a list of spaces that will be process. If empty all spaces are processed.
        :return:
        '''
        load_permission_sucess = self.get_space_permissions(load_space_list=space_list)
        if len(self.space_permission_list_raw)  > 0:
            for space in self.space_permission_list_raw:
                space_key = space['key']
                if space_key not in self.space_dict:
                    space_detail = {'id':space['id'],
                                    'key': space_key,
                                    'name': space['name'],
                                    'type': space['type']
                                    }
                    self.space_dict[space_key] = space_detail
                space_permission = space['permissions']
                for permission in space_permission:
                    if 'subjects' in permission:
                        if 'user' in permission['subjects']:
                            user_defined = False
                            account_details = permission['subjects']['user']['results'][0]
                            if account_details['accountType'] == 'atlassian':
                                user_id = account_details['accountId']
                                user_type = 'singel_user'
                                user_details = {'user_id': user_id,
                                                'email': account_details['email'],
                                                'name': account_details['publicName'],
                                                'type': user_type
                                                }
                                user_defined = True
                            # elif account_details['accountType'] == 'app':
                            #     user_type = 'app'
                            #     user_id = account_details['accountId']
                            #     user_details = {'user_id': user_id,
                            #                     'name': account_details['publicName'],
                            #                     'type': user_type
                            #                     }
                            #     user_defined = True
                        elif 'group' in permission['subjects']:
                            account_details = permission['subjects']['group']['results'][0]
                            user_type = 'group'
                            user_id = account_details['id']
                            user_details = {'user_id': user_id,
                                            'name': account_details['name'],
                                            'type': user_type
                                            }
                            user_defined = True
                        if user_defined:
                            self.space_user_dict[user_id] = user_details
                    if 'operation' in permission and user_defined:
                        operation = permission['operation']['operation']
                        target = permission['operation']['targetType']
                        permissions_type = find_permission_type(operation, target)
                        permission_detail = {'space_key': space_key,
                                             'user_id': user_id,
                                             'user_type': user_type,
                                             'operation': operation,
                                             'target': target,
                                             'permissions_type': permissions_type,
                                             'permissions_id':permission['id']
                                             }
                        self.space_permission_list.append(permission_detail)


    def write_to_excel(self, space_list=[]):
        '''
        Transforms the contence of space_permission list to an excel sheet where all spaces, users and theeir permission typs are listed .
        The sheet can be used for getiing an much needed overview of all confluence permission on a site.
        :param space_list: a list of spaces that is to be listed in the excel sheet. if empty then all spaces are listed.
        :output: The excel sheet "permission.xlsx"
        '''
        self.decode_space_persmissions(space_list=space_list)
        wb = openpyxl.Workbook()
        collumn_mapping = exel_collumn_mapping()
        ws = wb.active
        row_number = 1
        # writing cullumn names
        for cullumn in collumn_mapping:
            ws.cell(row=row_number, column=collumn_mapping[cullumn], value=cullumn)
        for space in self.space_dict:
            print(space)
            for user in self.space_user_dict:
                user_name = self.space_user_dict[user]['name']
                user_info_written = False
                for permission in self.space_permission_list:
                    if permission['space_key'] == space and permission['user_id'] == user:
                        if not user_info_written:
                            row_number += 1
                            ws.cell(row=row_number, column=collumn_mapping['space_key'], value=permission['space_key'])
                            ws.cell(row=row_number, column=collumn_mapping['user_name'], value=user_name)
                            ws.cell(row=row_number, column=collumn_mapping['user_id'], value=permission['user_id'])
                            ws.cell(row=row_number, column=collumn_mapping['user_type'], value=permission['user_type'])
                            user_info_written = True
                        if permission['permissions_type'] != 'unknown':
                            permission_name =  permission['permissions_type']
                            ws.cell(row=row_number, column=collumn_mapping[permission_name], value='X')
        wb.save('permission.xlsx')

    def bach_delete_user_permissions(self, space_list=[]):
        '''
        Delete all permision for all users and groups that are listed in the file "delete_users.csv" in all spaces
        listed in the file "confluence_space.csv"- If there are space listed in te parameter space_list one thoses are listed.
        :param space_list: a list of spaces that will be process. If empty all spaces are processed.
        :return:
        '''
        self.decode_space_persmissions(space_list=space_list)
        user_list = readFromCsv('delete_users.csv')
        for user in user_list:
            user_id = user['user_id']
            user_name = user['user_name']
            for space in self.space_dict:
                # removes all permissions for this space for this user exept view space
                for permission in self.space_permission_list:
                    permission_type =  permission['permissions_type']
                    if user_id == permission['user_id'] and permission['space_key'] == space:
                        if permission_type != 'view_space':
                            result = self.delete_space_permission(space,permission['permissions_id'])
                            if result == 204:
                                log_text = f'{permission_type} Permission for user {user_name} in space {space} has been deleted'
                                print(log_text)
                                logger.info(log_text)
                            else:
                                log_text = f'Delition of permission {permission_type} for user {user_name} in space {space} failed'
                                print(log_text)
                                logger.info (log_text)
                # removes view space for this user for this space
                for permission in self.                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                          space_permission_list:
                    permission_type =  permission['permissions_type']
                    if user_id == permission['user_id'] and permission['space_key'] == space:
                        if permission_type == 'view_space':
                            result = self.delete_space_permission(space,permission['permissions_id'])
                            if result == 204:
                                log_text = f'{permission_type} Permission for user/group {user_name} in space {space} has been deleted'
                                print(log_text)
                                logger.info(log_text)
                            else:
                                log_text = f'Delition of permission {permission_type} for user/group {user_name} in space {space} failed'
                                print(log_text)
                                logger.info(log_text)

    def replace_user_groups(self,usergrp_new, usergrp_old, space_list = []):
        '''

        :param usergrp_new: The name of the new user group. The goroup that will get the same permission as usergrp_old
        :param usergrp_old: The user group that will loose its permissions.
        :param space_list: a list of spaces that will be process. If empty all spaces are processed.
        :return: None
        '''
        self.decode_space_persmissions(space_list=space_list)
        for group in self.space_user_dict:
            if self.space_user_dict[group]['name'] == usergrp_old:
                id_usergrp_old = self.space_user_dict[group]['user_id']
                break
        for space in self.space_dict:
            view_space_not_set = True
            view_space_permission_found = False
            for permission in self.space_permission_list:
                if permission['space_key'] == space and permission['user_id'] == id_usergrp_old:
                    # creates view space permission for the new user group
                    if view_space_not_set:
                        status = self.set_space_permission(space,usergrp_new,'group','read', 'space')
                        view_space_not_set = False
                        if status in (200,204):
                            logtext = f'View space permission set for {usergrp_new} in space {space}'
                            print(logtext)
                            logger.info(logtext)
                        else:
                            logtext = f' Failed to set View space permission for {usergrp_new} in space {space}'
                            print(logtext)
                            logger.critical(logtext)
                            raise Exception('Insert of permission failed. Abort script')

                    # store the id for the view space permission so that it can be deleted last
                    if permission['permissions_type'] == 'view_space':
                        view_space_permission_id = permission['permissions_id']
                        view_space_permission_found = True
                    else:
                        # sets permission for new group
                        result = self.set_space_permission(space,usergrp_new,'group',permission['operation'], permission['target'])
                        if result in (200, 204):
                            logtext = f'Permission {permission["permissions_type"]} set for {usergrp_new} in space {space}'
                            print(logtext)
                            logger.info(logtext)
                        else:
                            logtext = f' Failed to set {permission["permissions_type"]} permission for {usergrp_new} in space {space}'
                            print(logtext)
                            logger.critical(logtext)
                            raise Exception('Insert of permission failed. Abort script')
                        # delete permission for old group
                        result = self.delete_space_permission(space,permission['permissions_id'])
                        if result in (200, 204):
                            logtext = f'Permission {permission["permissions_type"]} deleted for {usergrp_old} in space {space}'
                            print(logtext)
                            logger.info(logtext)
                        else:
                            logtext = f' Failed to set {permission["permissions_type"]} permission for {usergrp_old} in space {space}'
                            print(logtext)
                            logger.critical(logtext)
                 # delete view space permission for old user group
                    if view_space_permission_found:
                        result = self.delete_space_permission(space,view_space_permission_id)
                        if result in (200, 204):
                            logtext = f'Permission view space deleted for {usergrp_old} in space {space}'
                            print(logtext)
                            logger.info(logtext)
                        else:
                            logtext = f' Failed to set view space permission for {usergrp_old} in space {space}'
                            print(logtext)
                            logger.critical(logtext)















